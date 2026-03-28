import sqlite3
import os
import pathlib
import secrets
import uuid
from datetime import date, datetime, timedelta
from contextlib import contextmanager
from typing import Optional

from fastapi import FastAPI, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from models import create_tables

# ─── App setup ────────────────────────────────────────────────────────────────
app = FastAPI(title="Squash ELO API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = os.environ.get("DB_PATH", "squash.db")

# ─── Admin auth ───────────────────────────────────────────────────────────────
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
_admin_tokens: set[str] = set()

class AdminLogin(BaseModel):
    password: str

def require_admin(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Admin required")
    if authorization[7:] not in _admin_tokens:
        raise HTTPException(401, "Invalid or expired token")

# ─── Database ─────────────────────────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=OFF")
    # Legacy tables (kept for safety during migration window)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS players (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            name    TEXT UNIQUE NOT NULL,
            created TEXT DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS matches (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            p1      TEXT NOT NULL,
            p2      TEXT NOT NULL,
            s1      INTEGER NOT NULL,
            s2      INTEGER NOT NULL,
            created TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS challenges (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            challenger      TEXT NOT NULL,
            challenged      TEXT NOT NULL,
            required_wins   INTEGER NOT NULL DEFAULT 2,
            issued_date     TEXT NOT NULL,
            deadline_date   TEXT,
            status          TEXT NOT NULL DEFAULT 'open',
            winner          TEXT,
            resolved_date   TEXT,
            created         TEXT DEFAULT (datetime('now'))
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_matches_unique
            ON matches(date, p1, p2, s1, s2, created);
    """)
    # New tables
    create_tables(conn)
    conn.close()


init_db()

# ─── ELO Engine ───────────────────────────────────────────────────────────────
def validate_score(s1: int, s2: int) -> bool:
    return max(s1, s2) >= 11 and abs(s1 - s2) >= 2

def get_k(elo: float, match_count: int) -> int:
    if elo >= 2000: return 10
    if match_count >= 30: return 20
    return 40

def calc_elo(r_a: float, r_b: float, won: bool, mc: int, s_a: int, s_b: int) -> float:
    expected = 1 / (1 + 10 ** ((r_b - r_a) / 400))
    k = get_k(r_a, mc)
    margin = abs(s_a - s_b)
    multiplier = 1 + margin / max(s_a, s_b)
    return round(r_a + k * multiplier * ((1 if won else 0) - expected), 1)

def compute_state(matches_rows, player_names):
    ratings = {n: 1000.0 for n in player_names}
    counts  = {n: 0      for n in player_names}
    stats   = {
        n: {"name": n, "elo": 1000.0, "matches": 0, "wins": 0,
            "losses": 0, "peak": 1000.0, "low": 1000.0, "history": [1000.0]}
        for n in player_names
    }
    out = []

    for m in matches_rows:
        mid, dt = m["id"], m["date"]
        p1, p2, s1, s2 = m["p1"], m["p2"], m["s1"], m["s2"]
        is_valid = (p1 != p2 and p1 in ratings and p2 in ratings
                    and validate_score(s1, s2))

        p1pre = ratings.get(p1, 1000.0)
        p2pre = ratings.get(p2, 1000.0)

        if not is_valid:
            out.append({"id": mid, "date": dt, "p1": p1, "p2": p2,
                        "s1": s1, "s2": s2, "valid": False,
                        "winner": None, "p1pre": p1pre, "p2pre": p2pre,
                        "p1post": None, "p2post": None})
            continue

        p1wins = s1 > s2
        p1post = calc_elo(p1pre, p2pre, p1wins, counts[p1], s1, s2)
        p2post = calc_elo(p2pre, p1pre, not p1wins, counts[p2], s2, s1)

        ratings[p1] = p1post
        ratings[p2] = p2post
        counts[p1] += 1
        counts[p2] += 1

        winner = p1 if p1wins else p2
        out.append({"id": mid, "date": dt, "p1": p1, "p2": p2,
                    "s1": s1, "s2": s2, "valid": True,
                    "winner": winner, "p1pre": p1pre, "p2pre": p2pre,
                    "p1post": p1post, "p2post": p2post})

        for pname, post in [(p1, p1post), (p2, p2post)]:
            s = stats[pname]
            s["matches"] += 1
            if winner == pname: s["wins"] += 1
            else:               s["losses"] += 1
            s["peak"] = max(s["peak"], post)
            s["low"]  = min(s["low"],  post)
            s["history"].append(post)
            s["elo"] = post

    return out, stats


# ─── Season helpers ───────────────────────────────────────────────────────────
_SEASON_NAMES = {"Q1": "Winter", "Q2": "Spring", "Q3": "Summer", "Q4": "Autumn"}
_SEASON_STARTS = {"Q1": (1, 1), "Q2": (4, 1), "Q3": (7, 1), "Q4": (10, 1)}
_SEASON_ENDS   = {"Q1": (3, 31), "Q2": (6, 30), "Q3": (9, 30), "Q4": (12, 31)}

def get_season_id(date_str: str) -> str:
    d = datetime.strptime(date_str, "%Y-%m-%d")
    q = (d.month - 1) // 3 + 1
    return f"{d.year}-Q{q}"

def get_season_name(sid: str) -> str:
    year, q = sid.split("-")
    return f"{_SEASON_NAMES[q]} {year}"

def get_season_dates(sid: str) -> tuple:
    year, q = sid.split("-")
    yr = int(year)
    s = _SEASON_STARTS[q]; e = _SEASON_ENDS[q]
    return f"{yr}-{s[0]:02d}-{s[1]:02d}", f"{yr}-{e[0]:02d}-{e[1]:02d}"

def current_season_id() -> str:
    return get_season_id(str(date.today()))


def compute_seasons(match_rows, player_names: list) -> dict:
    """Return per-season ELO standings with soft resets between seasons."""
    # Group matches by season
    seasons_matches: dict = {}
    for m in match_rows:
        sid = get_season_id(m["date"])
        seasons_matches.setdefault(sid, []).append(m)

    # Ensure current season is always present
    cur_sid = current_season_id()
    seasons_matches.setdefault(cur_sid, [])

    # Sort seasons chronologically
    sorted_sids = sorted(seasons_matches.keys())

    # Carry-over ELOs between seasons (starts at 1000 for new players)
    carry: dict = {}  # name → elo at end of previous season

    seasons_out = {}
    for sid in sorted_sids:
        s_matches = seasons_matches[sid]

        # Compute season-start ELOs (after soft reset)
        season_start: dict = {}
        for name in player_names:
            prev = carry.get(name)
            if prev is None:
                season_start[name] = 1000.0
            else:
                season_start[name] = round(prev + (1200 - prev) * 0.3, 1)

        # Run ELO compute with season-start values
        # We need to seed ratings with season_start ELOs; reuse compute_state logic inline
        ratings = dict(season_start)
        counts  = {n: 0 for n in player_names}
        player_stats = {
            n: {"name": n, "elo": season_start[n], "season_start_elo": season_start[n],
                "wins": 0, "losses": 0, "matches": 0, "peak": season_start[n], "low": season_start[n]}
            for n in player_names
        }

        # Track streak per player (list of W/L for this season)
        streak_seq: dict = {n: [] for n in player_names}
        # Track giant killer (wins vs higher pre-match ELO)
        giant_kills: dict = {n: 0 for n in player_names}
        # Track peak elo within season
        computed_season_matches = []

        for m in s_matches:
            p1, p2, s1, s2 = m["p1"], m["p2"], m["s1"], m["s2"]
            is_valid = (p1 != p2 and p1 in ratings and p2 in ratings
                        and validate_score(s1, s2))
            if not is_valid:
                computed_season_matches.append({"id": m["id"], "valid": False,
                    "p1": p1, "p2": p2, "s1": s1, "s2": s2,
                    "p1pre": ratings.get(p1, 1000.0), "p2pre": ratings.get(p2, 1000.0),
                    "winner": None})
                continue

            p1pre = ratings[p1]; p2pre = ratings[p2]
            p1wins = s1 > s2
            p1post = calc_elo(p1pre, p2pre, p1wins,  counts[p1], s1, s2)
            p2post = calc_elo(p2pre, p1pre, not p1wins, counts[p2], s2, s1)

            winner = p1 if p1wins else p2
            loser  = p2 if p1wins else p1

            computed_season_matches.append({"id": m["id"], "valid": True,
                "p1": p1, "p2": p2, "s1": s1, "s2": s2,
                "p1pre": p1pre, "p2pre": p2pre, "winner": winner})

            # Giant killer: winner beat someone with higher pre-match ELO
            winner_pre = p1pre if p1wins else p2pre
            loser_pre  = p2pre if p1wins else p1pre
            if loser_pre > winner_pre:
                giant_kills[winner] += 1

            for pname, post, won in [(p1, p1post, p1wins), (p2, p2post, not p1wins)]:
                ratings[pname] = post
                counts[pname] += 1
                st = player_stats[pname]
                st["matches"] += 1
                st["elo"] = post
                st["peak"] = max(st["peak"], post)
                st["low"]  = min(st["low"],  post)
                if won: st["wins"] += 1
                else:   st["losses"] += 1
                streak_seq[pname].append("W" if won else "L")

        # Compute streak for each player (current run at end of season)
        for name, seq in streak_seq.items():
            s = 0
            if seq:
                last = seq[-1]
                for r in reversed(seq):
                    if r == last: s += 1
                    else: break
                player_stats[name]["streak"] = s if last == "W" else -s
            else:
                player_stats[name]["streak"] = 0
            player_stats[name]["giant_kills"] = giant_kills[name]
            player_stats[name]["delta"] = round(
                player_stats[name]["elo"] - player_stats[name]["season_start_elo"], 1)

        # Carry ELOs forward
        for name in player_names:
            carry[name] = player_stats[name]["elo"]

        start_date, end_date = get_season_dates(sid)
        is_current = (sid == cur_sid)
        today = date.today()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        days_remaining = max(0, (end_dt - today).days) if is_current else 0

        seasons_out[sid] = {
            "season_id": sid,
            "name": get_season_name(sid),
            "start": start_date,
            "end": end_date,
            "is_current": is_current,
            "match_count": sum(1 for m in computed_season_matches if m["valid"]),
            "days_remaining": days_remaining,
            "players": player_stats,
            "matches": computed_season_matches,
        }

    return seasons_out


def compute_awards(season: dict, min_matches: int = 3) -> dict:
    players = season["players"]
    qualified = [p for p in players.values() if p["matches"] >= min_matches]

    champion = max(players.values(), key=lambda p: p["elo"]) if players else None
    most_improved = max(qualified, key=lambda p: p["delta"]) if qualified else None
    most_active   = max(players.values(), key=lambda p: p["matches"]) if players else None
    most_active   = most_active if most_active and most_active["matches"] >= 1 else None
    giant_killer_candidates = [p for p in qualified if p["giant_kills"] > 0]
    giant_killer  = max(giant_killer_candidates, key=lambda p: p["giant_kills"]) if giant_killer_candidates else None

    def fmt(p, key):
        if p is None: return None
        if key == "champion":
            return {"name": p["name"], "detail": f"{p['elo']:.0f} ELO"}
        if key == "most_improved":
            d = p["delta"]; sign = "+" if d >= 0 else ""
            return {"name": p["name"], "detail": f"{sign}{d:.0f} ELO from season start"}
        if key == "most_active":
            return {"name": p["name"], "detail": f"{p['matches']} matches played"}
        if key == "giant_killer":
            return {"name": p["name"], "detail": f"{p['giant_kills']} wins vs. higher-ranked"}

    return {
        "champion":      fmt(champion, "champion"),
        "most_improved": fmt(most_improved, "most_improved"),
        "most_active":   fmt(most_active, "most_active"),
        "giant_killer":  fmt(giant_killer, "giant_killer"),
    }


# ─── New-schema helpers ────────────────────────────────────────────────────────
def new_uuid():
    return str(uuid.uuid4())

def get_main_league(db):
    row = db.execute("SELECT id FROM league WHERE name='Main League' LIMIT 1").fetchone()
    return row["id"] if row else None

def get_system_player(db):
    row = db.execute("SELECT id FROM player WHERE display_name='System' LIMIT 1").fetchone()
    return row["id"] if row else None

def get_match_rows_new(db):
    return db.execute("""
        SELECT
            COALESCE(m.legacy_id, 900000000 + CAST(m.rowid AS INTEGER)) AS id,
            m.played_at     AS date,
            pa.display_name AS p1,
            pb.display_name AS p2,
            ms.player_a_score AS s1,
            ms.player_b_score AS s2,
            m.created_at    AS created
        FROM match m
        JOIN player pa ON pa.id = m.player_a_id
        JOIN player pb ON pb.id = m.player_b_id
        JOIN match_set ms ON ms.match_id = m.id AND ms.set_number = 1
        ORDER BY COALESCE(m.legacy_id, 999999999) ASC, m.created_at ASC
    """).fetchall()

def get_player_names_new(db):
    rows = db.execute(
        "SELECT display_name FROM player WHERE display_name != 'System' ORDER BY display_name"
    ).fetchall()
    return [r["display_name"] for r in rows]

def _update_membership_stats(db, league_id, player_names, computed_matches):
    """Recompute and persist league_membership stats for all players after any match change."""
    streak_seq    = {name: [] for name in player_names}
    ch_tracker    = {name: 1000.0 for name in player_names}
    ch_at_tracker = {name: None for name in player_names}
    rating_now    = {name: 1000.0 for name in player_names}
    mp            = {name: 0 for name in player_names}
    mw            = {name: 0 for name in player_names}
    last_played   = {name: None for name in player_names}

    for cm in computed_matches:
        if not cm["valid"]:
            continue
        p1, p2 = cm["p1"], cm["p2"]
        if p1 not in rating_now or p2 not in rating_now:
            continue
        winner = cm["winner"]
        for pname, post in [(p1, cm["p1post"]), (p2, cm["p2post"])]:
            won = (winner == pname)
            mp[pname] += 1
            if won:
                mw[pname] += 1
            rating_now[pname] = post
            if post > ch_tracker[pname]:
                ch_tracker[pname]    = post
                ch_at_tracker[pname] = cm["date"]
            streak_seq[pname].append("W" if won else "L")
            last_played[pname] = cm["date"]

    now_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")

    for name in player_names:
        seq = streak_seq[name]
        # current streak
        if seq:
            last = seq[-1]
            cur = 0
            for r in reversed(seq):
                if r == last:
                    cur += 1
                else:
                    break
            current_streak = cur if last == "W" else -cur
        else:
            current_streak = 0

        # longest win streak
        longest = 0
        run = 0
        for r in seq:
            if r == "W":
                run += 1
                longest = max(longest, run)
            else:
                run = 0

        ch_at = (ch_at_tracker[name] + "T00:00:00"
                 if ch_at_tracker[name] and "T" not in ch_at_tracker[name]
                 else ch_at_tracker[name])
        lp = last_played[name]
        rating_updated_at = (lp + "T00:00:00" if lp and "T" not in lp else lp)

        db.execute(
            """UPDATE league_membership SET
               rating=?, career_high=?, career_high_at=?,
               current_streak=?, longest_win_streak=?,
               matches_played=?, matches_won=?,
               rating_updated_at=?
               WHERE player_id=(SELECT id FROM player WHERE display_name=?)
               AND league_id=?""",
            (rating_now[name], ch_tracker[name], ch_at,
             current_streak, longest,
             mp[name], mw[name],
             rating_updated_at,
             name, league_id)
        )


# ─── Models ───────────────────────────────────────────────────────────────────
class NewMatch(BaseModel):
    date: Optional[str] = None
    p1: str
    p2: str
    s1: int
    s2: int

class NewPlayer(BaseModel):
    name: str

class ChallengeCreate(BaseModel):
    challenger: str
    challenged: str
    required_wins: int
    deadline_days: Optional[int] = None

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.get("/api")
def root():
    return {"status": "ok", "service": "Squash ELO API"}


@app.get("/api/players")
def get_players():
    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)
    _, stats = compute_state(match_rows, names)
    return sorted(stats.values(), key=lambda x: -x["elo"])


@app.post("/api/players", status_code=201)
def add_player(body: NewPlayer):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Name cannot be empty")
    with get_db() as db:
        if db.execute("SELECT id FROM player WHERE display_name=?", (name,)).fetchone():
            raise HTTPException(409, f"Player '{name}' already exists")
        main_league_id = get_main_league(db)
        if not main_league_id:
            raise HTTPException(500, "Main League not found — run migration first")
        email = f"{name.lower().replace(' ', '.')}@placeholder.local"
        now_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
        pid = new_uuid()
        db.execute(
            "INSERT INTO player (id, display_name, avatar_url, email, created_at, last_active_at) VALUES (?,?,?,?,?,?)",
            (pid, name, None, email, now_iso, now_iso)
        )
        lm_id = new_uuid()
        db.execute(
            """INSERT INTO league_membership
               (id, player_id, league_id, role, joined_at, is_active,
                rating, rating_deviation, volatility, rating_updated_at,
                career_high, career_high_at, current_streak, longest_win_streak,
                matches_played, matches_won)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (lm_id, pid, main_league_id, "member", now_iso, 1,
             1000.0, 350.0, 0.06, None,
             1000.0, None, 0, 0, 0, 0)
        )
    return {"name": name}


@app.get("/api/matches")
def get_matches():
    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)
    computed, _ = compute_state(match_rows, names)
    return computed


@app.post("/api/matches", status_code=201)
def log_match(body: NewMatch):
    if body.p1 == body.p2:
        raise HTTPException(400, "Players must be different")
    if not validate_score(body.s1, body.s2):
        raise HTTPException(400, "Invalid score — win by 2 with ≥11 pts (e.g. 11-7, 12-10, 15-13)")

    match_date = body.date or str(date.today())
    try:
        datetime.strptime(match_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date — use YYYY-MM-DD")

    with get_db() as db:
        for pname in [body.p1, body.p2]:
            if not db.execute("SELECT id FROM player WHERE display_name=?", (pname,)).fetchone():
                raise HTTPException(404, f"Player '{pname}' not found")

        main_league_id = get_main_league(db)
        if not main_league_id:
            raise HTTPException(500, "Main League not found — run migration first")
        system_player_id = get_system_player(db)

        # Assign legacy_id for API compatibility
        legacy_id_row = db.execute("SELECT COALESCE(MAX(legacy_id),0)+1 FROM match").fetchone()
        legacy_id = legacy_id_row[0]

        p1_row = db.execute("SELECT id FROM player WHERE display_name=?", (body.p1,)).fetchone()
        p2_row = db.execute("SELECT id FROM player WHERE display_name=?", (body.p2,)).fetchone()
        player_a_id = p1_row["id"]
        player_b_id = p2_row["id"]

        # Determine winner
        winner_id = player_a_id if body.s1 > body.s2 else player_b_id

        now_iso  = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
        match_uuid = new_uuid()

        db.execute(
            """INSERT INTO match
               (id, league_id, player_a_id, player_b_id, winner_id,
                match_type, status, logged_by, confirmed_by, confirmed_at,
                venue, notes, played_at, created_at, legacy_id)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (match_uuid, main_league_id, player_a_id, player_b_id, winner_id,
             "league", "confirmed", system_player_id or match_uuid, None, None,
             None, None, match_date, now_iso, legacy_id)
        )

        set_uuid = new_uuid()
        db.execute(
            """INSERT INTO match_set
               (id, match_id, set_number, player_a_score, player_b_score, winner_id)
               VALUES (?,?,?,?,?,?)""",
            (set_uuid, match_uuid, 1, body.s1, body.s2, winner_id)
        )

        # Fetch all rows and recompute
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)

        computed, _ = compute_state(match_rows, names)
        _update_membership_stats(db, main_league_id, names, computed)

    match_result = next((m for m in computed if m["id"] == legacy_id), None)
    if match_result is None:
        raise HTTPException(status_code=500, detail="Match was saved but could not be retrieved")
    return match_result


@app.post("/api/admin/login")
def admin_login(body: AdminLogin):
    if not ADMIN_PASSWORD or body.password != ADMIN_PASSWORD:
        raise HTTPException(401, "Invalid password")
    token = secrets.token_hex(32)
    _admin_tokens.add(token)
    return {"token": token}


@app.post("/api/reset", status_code=200)
def reset_db(authorization: Optional[str] = Header(None)):
    require_admin(authorization)
    with get_db() as db:
        db.execute("PRAGMA foreign_keys=OFF")
        # New tables — delete in dependency order
        db.execute("DELETE FROM rating_snapshot")
        db.execute("DELETE FROM match_set")
        db.execute("DELETE FROM challenge")
        db.execute("DELETE FROM match")
        db.execute("DELETE FROM league_membership")
        db.execute("DELETE FROM league")
        db.execute("DELETE FROM player")
        # Legacy tables
        db.execute("DELETE FROM matches")
        db.execute("DELETE FROM players")
        db.execute("DELETE FROM challenges")
        db.execute("DELETE FROM sqlite_sequence WHERE name IN ('matches','players','challenges')")

        # Re-seed System player + Main League
        now_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
        system_id = new_uuid()
        db.execute(
            "INSERT INTO player (id, display_name, avatar_url, email, created_at, last_active_at) VALUES (?,?,?,?,?,?)",
            (system_id, "System", None, "system@placeholder.local", now_iso, now_iso)
        )
        league_id  = new_uuid()
        invite_code = uuid.uuid4().hex[:8].upper()
        db.execute(
            """INSERT INTO league
               (id, name, created_by, league_type, scoring_format, points_to_win,
                is_active, season_start, season_end, invite_code, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (league_id, "Main League", system_id, "ladder", "single_game",
             11, 1, None, None, invite_code, now_iso)
        )
        db.execute("PRAGMA foreign_keys=ON")
    return {"status": "reset"}


@app.delete("/api/matches/{match_id}", status_code=200)
def delete_match(match_id: int, authorization: Optional[str] = Header(None)):
    require_admin(authorization)
    with get_db() as db:
        row = db.execute("SELECT id FROM match WHERE legacy_id=?", (match_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Match not found")
        match_uuid = row["id"]

        db.execute("DELETE FROM match_set WHERE match_id=?", (match_uuid,))
        db.execute("DELETE FROM rating_snapshot WHERE match_id=?", (match_uuid,))
        db.execute("DELETE FROM match WHERE id=?", (match_uuid,))

        # Recompute all membership ratings
        main_league_id = get_main_league(db)
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)
        computed, _ = compute_state(match_rows, names)
        if main_league_id:
            _update_membership_stats(db, main_league_id, names, computed)

    return {"deleted": match_id}


@app.get("/api/stats")
def get_stats():
    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)
    computed, stats = compute_state(match_rows, names)
    return {
        "players": sorted(stats.values(), key=lambda x: -x["elo"]),
        "matches": computed,
    }


@app.get("/api/export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)

    computed, stats = compute_state(match_rows, names)

    wb   = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="1B3A6B")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri")

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")

    # Leaderboard sheet
    ws = wb.active
    ws.title = "Leaderboard"
    style_header(ws, ["Rank","Player","ELO","Matches","Wins","Losses","Win %","Peak ELO","Low ELO"])
    for rank, p in enumerate(sorted(stats.values(), key=lambda x: -x["elo"]), 1):
        ws.append([rank, p["name"], p["elo"], p["matches"], p["wins"], p["losses"],
                   round(p["wins"] / p["matches"], 3) if p["matches"] else 0,
                   p["peak"], p["low"]])

    # Match Log sheet
    ws2 = wb.create_sheet("Match Log")
    style_header(ws2, ["ID","Date","Player 1","Player 2","P1 Score","P2 Score",
                        "Valid","Winner","P1 Pre-ELO","P2 Pre-ELO","P1 Post-ELO","P2 Post-ELO"])
    for m in computed:
        ws2.append([m["id"], m["date"], m["p1"], m["p2"], m["s1"], m["s2"],
                    "VALID" if m["valid"] else "INVALID", m["winner"] or "",
                    m["p1pre"], m["p2pre"],
                    m["p1post"] if m["p1post"] is not None else "",
                    m["p2post"] if m["p2post"] is not None else ""])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=squash_elo.xlsx"}
    )


@app.get("/api/seasons")
def get_seasons():
    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)

    if not names:
        return []

    seasons = compute_seasons(match_rows, names)

    result = []
    for sid in sorted(seasons.keys(), reverse=True):
        s = seasons[sid]
        players_sorted = sorted(s["players"].values(), key=lambda p: -p["elo"])
        champion = players_sorted[0]["name"] if players_sorted and not s["is_current"] else None
        result.append({
            "season_id":    sid,
            "name":         s["name"],
            "start":        s["start"],
            "end":          s["end"],
            "is_current":   s["is_current"],
            "match_count":  s["match_count"],
            "champion":     champion,
        })

    return result


@app.get("/api/seasons/{season_id}")
def get_season(season_id: str):
    with get_db() as db:
        match_rows = get_match_rows_new(db)
        names      = get_player_names_new(db)

    if not names:
        raise HTTPException(404, "No players found")

    seasons = compute_seasons(match_rows, names)
    if season_id not in seasons:
        raise HTTPException(404, f"Season '{season_id}' not found")

    s = seasons[season_id]
    standings = sorted(s["players"].values(), key=lambda p: -p["elo"])
    for i, p in enumerate(standings):
        p["rank"] = i + 1

    return {
        "season_id":     s["season_id"],
        "name":          s["name"],
        "start":         s["start"],
        "end":           s["end"],
        "is_current":    s["is_current"],
        "days_remaining": s["days_remaining"],
        "match_count":   s["match_count"],
        "standings":     standings,
        "awards":        compute_awards(s),
    }


# ─── Challenger Mode ──────────────────────────────────────────────────────────
@app.post("/api/challenges", status_code=201)
def create_challenge(body: ChallengeCreate):
    if not body.challenger.strip() or not body.challenged.strip():
        raise HTTPException(422, "challenger and challenged must be non-empty strings")
    if body.challenger == body.challenged:
        raise HTTPException(400, "A player cannot challenge themselves")
    if body.required_wins not in (2, 3):
        raise HTTPException(400, "required_wins must be 2 or 3")
    if body.deadline_days is not None and body.deadline_days not in (7, 14, 30):
        raise HTTPException(400, "deadline_days must be 7, 14, 30, or null")

    issued_date = datetime.utcnow().strftime("%Y-%m-%d")
    deadline_date = (
        (datetime.utcnow() + timedelta(days=body.deadline_days)).strftime("%Y-%m-%d")
        if body.deadline_days is not None else None
    )
    expires_at = (deadline_date + "T23:59:59") if deadline_date else None
    now_iso    = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")

    with get_db() as db:
        for pname in [body.challenger, body.challenged]:
            if not db.execute("SELECT id FROM player WHERE display_name=?", (pname,)).fetchone():
                raise HTTPException(404, f"Player not found: {pname}")

        challenger_row = db.execute("SELECT id FROM player WHERE display_name=?", (body.challenger,)).fetchone()
        challenged_row = db.execute("SELECT id FROM player WHERE display_name=?", (body.challenged,)).fetchone()
        challenger_uuid = challenger_row["id"]
        challenged_uuid = challenged_row["id"]

        main_league_id = get_main_league(db)
        if not main_league_id:
            raise HTTPException(500, "Main League not found — run migration first")

        # No existing pending challenge between same pair
        existing = db.execute(
            """SELECT id FROM challenge
               WHERE status = 'pending'
               AND ((challenger_id = ? AND challenged_id = ?)
                    OR (challenger_id = ? AND challenged_id = ?))""",
            (challenger_uuid, challenged_uuid, challenged_uuid, challenger_uuid)
        ).fetchone()
        if existing:
            raise HTTPException(400, "An open challenge already exists between these players")

        legacy_id_row = db.execute("SELECT COALESCE(MAX(legacy_id),0)+1 FROM challenge").fetchone()
        legacy_id = legacy_id_row[0]

        ch_uuid = new_uuid()
        db.execute(
            """INSERT INTO challenge
               (id, league_id, challenger_id, challenged_id,
                status, match_id, message, expires_at,
                created_at, responded_at,
                legacy_id, legacy_required_wins, legacy_winner, legacy_resolved_date)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (ch_uuid, main_league_id, challenger_uuid, challenged_uuid,
             "pending", None, None, expires_at,
             now_iso, None,
             legacy_id, body.required_wins, None, None)
        )

    return {
        "id":            legacy_id,
        "challenger":    body.challenger,
        "challenged":    body.challenged,
        "required_wins": body.required_wins,
        "issued_date":   issued_date,
        "deadline_date": deadline_date,
        "status":        "open",
        "winner":        None,
        "resolved_date": None,
        "created":       now_iso,
    }


@app.get("/api/challenges")
def get_challenges():
    today = datetime.utcnow().strftime("%Y-%m-%d")

    with get_db() as db:
        rows = db.execute(
            """SELECT
                   c.id AS uuid,
                   c.legacy_id,
                   c.legacy_required_wins,
                   c.legacy_winner,
                   c.legacy_resolved_date,
                   c.status,
                   c.expires_at,
                   c.created_at,
                   c.responded_at,
                   pa.display_name AS challenger,
                   pb.display_name AS challenged
               FROM challenge c
               JOIN player pa ON pa.id = c.challenger_id
               JOIN player pb ON pb.id = c.challenged_id
               ORDER BY c.created_at DESC"""
        ).fetchall()

        # Status translation: DB vocab → API vocab
        _status_to_api = {
            "pending":  "open",
            "accepted": "resolved",
            "expired":  "expired",
            "declined": "cancelled",
        }

        result = []
        for row in rows:
            ch = dict(row)
            challenger    = ch["challenger"]
            challenged    = ch["challenged"]
            required_wins = ch["legacy_required_wins"]
            stored_status = ch["status"]  # DB vocab
            expires_at    = ch["expires_at"]
            deadline_date = expires_at[:10] if expires_at else None
            issued_date   = ch["created_at"][:10] if ch["created_at"] else None

            # Step 1 — fetch relevant matches since issued_date
            match_rows = db.execute(
                """SELECT pa.display_name AS p1, pb.display_name AS p2,
                          ms.player_a_score AS s1, ms.player_b_score AS s2,
                          m.played_at AS date
                   FROM match m
                   JOIN player pa ON pa.id = m.player_a_id
                   JOIN player pb ON pb.id = m.player_b_id
                   JOIN match_set ms ON ms.match_id = m.id AND ms.set_number = 1
                   WHERE ((pa.display_name = ? AND pb.display_name = ?)
                       OR (pa.display_name = ? AND pb.display_name = ?))
                   AND m.played_at >= ?
                   ORDER BY m.played_at ASC, m.created_at ASC""",
                (challenger, challenged, challenged, challenger, issued_date)
            ).fetchall()

            # Step 2 — walk matches, stop when series decided
            challenger_wins = 0
            challenged_wins = 0
            computed_winner = None
            computed_resolved_date = None

            for m in match_rows:
                p1, p2, s1, s2 = m["p1"], m["p2"], m["s1"], m["s2"]
                match_winner = p1 if s1 > s2 else p2
                if match_winner == challenger:
                    challenger_wins += 1
                else:
                    challenged_wins += 1

                if challenger_wins >= required_wins:
                    computed_winner = challenger
                    computed_resolved_date = m["date"][:10]
                    break
                if challenged_wins >= required_wins:
                    computed_winner = challenged
                    computed_resolved_date = m["date"][:10]
                    break

            # Step 3 — compute status (in DB vocab)
            if stored_status == "declined":
                new_db_status = "declined"
                computed_winner = ch["legacy_winner"]
                computed_resolved_date = ch["legacy_resolved_date"]
            elif computed_winner:
                new_db_status = "accepted"
            elif deadline_date and today > deadline_date:
                new_db_status = "expired"
            else:
                new_db_status = "pending"

            # Step 4 — write-back if status changed
            if (new_db_status != stored_status
                    or computed_winner != ch.get("legacy_winner")
                    or computed_resolved_date != ch.get("legacy_resolved_date")):
                responded_at = (computed_resolved_date + "T00:00:00"
                                if computed_resolved_date else ch["responded_at"])
                db.execute(
                    """UPDATE challenge SET
                       status=?, legacy_winner=?, legacy_resolved_date=?, responded_at=?
                       WHERE id=?""",
                    (new_db_status, computed_winner, computed_resolved_date,
                     responded_at, ch["uuid"])
                )

            # Step 5 — days_remaining
            if deadline_date:
                deadline_dt = datetime.strptime(deadline_date, "%Y-%m-%d")
                today_dt    = datetime.strptime(today, "%Y-%m-%d")
                days_remaining = max(0, (deadline_dt - today_dt).days)
            else:
                days_remaining = None

            # Return in old API shape
            result.append({
                "id":              ch["legacy_id"],
                "challenger":      challenger,
                "challenged":      challenged,
                "required_wins":   required_wins,
                "issued_date":     issued_date,
                "deadline_date":   deadline_date,
                "status":          _status_to_api.get(new_db_status, new_db_status),
                "winner":          computed_winner,
                "resolved_date":   computed_resolved_date,
                "created":         ch["created_at"],
                "challenger_wins": challenger_wins,
                "challenged_wins": challenged_wins,
                "days_remaining":  days_remaining,
            })

    return result


@app.delete("/api/challenges/{challenge_id}", status_code=200)
def cancel_challenge(challenge_id: int, authorization: Optional[str] = Header(None)):
    require_admin(authorization)
    with get_db() as db:
        row = db.execute(
            "SELECT id, status FROM challenge WHERE legacy_id=?", (challenge_id,)
        ).fetchone()
        if not row:
            raise HTTPException(404, "Challenge not found")
        status = row["status"]
        if status in ("accepted", "declined"):
            api_status = "resolved" if status == "accepted" else "cancelled"
            raise HTTPException(400, f"Challenge is already {api_status} and cannot be cancelled")
        db.execute(
            "UPDATE challenge SET status='declined', responded_at=datetime('now') WHERE id=?",
            (row["id"],)
        )
    return {"message": "Challenge cancelled"}


# ─── Serve frontend static files (must come after all API routes) ────────────
STATIC_DIR = pathlib.Path(__file__).parent / "static"
if STATIC_DIR.exists():
    app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")
