import sqlite3
import os
from datetime import date
from contextlib import contextmanager
from typing import Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

# ─── App setup ────────────────────────────────────────────────────────────────
app = FastAPI(title="Squash ELO API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = os.environ.get("DB_PATH", "squash.db")

# ─── Database ─────────────────────────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    # Use a plain connection for init so executescript doesn't fight the context manager
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS players (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                name    TEXT UNIQUE NOT NULL,
                created TEXT DEFAULT (date('now'))
            );

            CREATE TABLE IF NOT EXISTS matches (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                date    TEXT NOT NULL,
                p1      TEXT NOT NULL,
                p2      TEXT NOT NULL,
                s1      INTEGER NOT NULL,
                s2      INTEGER NOT NULL,
                created TEXT DEFAULT (datetime('now'))
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_matches_unique
                ON matches(date, p1, p2, s1, s2, created);
        """)

        # Seed players (safe — UNIQUE constraint on name)
        for name in ["Alice","Bob","Charlie","Diana","Eve",
                     "Frank","Grace","Henry","Iris","Jack"]:
            conn.execute("INSERT OR IGNORE INTO players (name) VALUES (?)", (name,))

        # Seed matches only if table is empty
        count = conn.execute("SELECT COUNT(*) FROM matches").fetchone()[0]
        if count == 0:
            seeds = [
                ("2024-01-10", "Alice",   "Bob",     11, 7),
                ("2024-01-10", "Charlie", "Alice",   11, 9),
                ("2024-01-11", "Bob",     "Charlie", 11, 13),
                ("2024-01-12", "Alice",   "Charlie", 11, 10),
                ("2024-01-13", "Bob",     "Alice",   15, 13),
            ]
            conn.executemany(
                "INSERT INTO matches (date,p1,p2,s1,s2) VALUES (?,?,?,?,?)", seeds
            )

        conn.commit()
    finally:
        conn.close()


init_db()

# ─── ELO Engine ───────────────────────────────────────────────────────────────
def validate_score(s1: int, s2: int) -> bool:
    return max(s1, s2) >= 11 and abs(s1 - s2) >= 2

def get_k(elo: float, match_count: int) -> int:
    if elo >= 2000: return 10
    if match_count >= 30: return 20
    return 40

def calc_elo(r_a: float, r_b: float, won: bool, mc: int) -> float:
    expected = 1 / (1 + 10 ** ((r_b - r_a) / 400))
    k = get_k(r_a, mc)
    return round(r_a + k * ((1 if won else 0) - expected), 1)

def compute_state(matches_rows, player_names):
    ratings = {n: 1000.0 for n in player_names}
    counts  = {n: 0      for n in player_names}
    stats   = {
        n: {"name": n, "elo": 1000.0, "matches": 0, "wins": 0,
            "losses": 0, "peak": 1000.0, "low": 1000.0, "history": [1000.0]}
        for n in player_names
    }
    out = []

    for m in matches_rows:
        mid, dt = m["id"], m["date"]
        p1, p2, s1, s2 = m["p1"], m["p2"], m["s1"], m["s2"]
        is_valid = (p1 != p2 and p1 in ratings and p2 in ratings
                    and validate_score(s1, s2))

        p1pre = ratings.get(p1, 1000.0)
        p2pre = ratings.get(p2, 1000.0)

        if not is_valid:
            out.append({"id": mid, "date": dt, "p1": p1, "p2": p2,
                        "s1": s1, "s2": s2, "valid": False,
                        "winner": None, "p1pre": p1pre, "p2pre": p2pre,
                        "p1post": None, "p2post": None})
            continue

        p1wins = s1 > s2
        p1post = calc_elo(p1pre, p2pre, p1wins, counts[p1])
        p2post = calc_elo(p2pre, p1pre, not p1wins, counts[p2])

        ratings[p1] = p1post
        ratings[p2] = p2post
        counts[p1] += 1
        counts[p2] += 1

        winner = p1 if p1wins else p2
        out.append({"id": mid, "date": dt, "p1": p1, "p2": p2,
                    "s1": s1, "s2": s2, "valid": True,
                    "winner": winner, "p1pre": p1pre, "p2pre": p2pre,
                    "p1post": p1post, "p2post": p2post})

        for pname, post in [(p1, p1post), (p2, p2post)]:
            s = stats[pname]
            s["matches"] += 1
            if winner == pname: s["wins"] += 1
            else:               s["losses"] += 1
            s["peak"] = max(s["peak"], post)
            s["low"]  = min(s["low"],  post)
            s["history"].append(post)
            s["elo"] = post

    return out, stats

# ─── Models ───────────────────────────────────────────────────────────────────
class NewMatch(BaseModel):
    date: Optional[str] = None
    p1: str
    p2: str
    s1: int
    s2: int

class NewPlayer(BaseModel):
    name: str

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.get("/")
def root():
    return {"status": "ok", "service": "Squash ELO API"}


@app.get("/players")
def get_players():
    with get_db() as db:
        rows       = db.execute("SELECT name FROM players ORDER BY name").fetchall()
        match_rows = db.execute("SELECT id,date,p1,p2,s1,s2 FROM matches ORDER BY id").fetchall()
    names = [r["name"] for r in rows]
    _, stats = compute_state(match_rows, names)
    return sorted(stats.values(), key=lambda x: -x["elo"])


@app.post("/players", status_code=201)
def add_player(body: NewPlayer):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Name cannot be empty")
    with get_db() as db:
        if db.execute("SELECT id FROM players WHERE name=?", (name,)).fetchone():
            raise HTTPException(409, f"Player '{name}' already exists")
        db.execute("INSERT INTO players (name) VALUES (?)", (name,))
    return {"name": name}


@app.get("/matches")
def get_matches():
    with get_db() as db:
        rows  = db.execute("SELECT id,date,p1,p2,s1,s2 FROM matches ORDER BY id").fetchall()
        names = [r["name"] for r in db.execute("SELECT name FROM players").fetchall()]
    computed, _ = compute_state(rows, names)
    return computed


@app.post("/matches", status_code=201)
def log_match(body: NewMatch):
    if body.p1 == body.p2:
        raise HTTPException(400, "Players must be different")
    if not validate_score(body.s1, body.s2):
        raise HTTPException(400, "Invalid score — win by 2 with ≥11 pts (e.g. 11-7, 12-10, 15-13)")

    match_date = body.date or str(date.today())

    with get_db() as db:
        for pname in [body.p1, body.p2]:
            if not db.execute("SELECT id FROM players WHERE name=?", (pname,)).fetchone():
                raise HTTPException(404, f"Player '{pname}' not found")

        cur = db.execute(
            "INSERT INTO matches (date,p1,p2,s1,s2) VALUES (?,?,?,?,?)",
            (match_date, body.p1, body.p2, body.s1, body.s2)
        )
        new_id = cur.lastrowid
        rows  = db.execute("SELECT id,date,p1,p2,s1,s2 FROM matches ORDER BY id").fetchall()
        names = [r["name"] for r in db.execute("SELECT name FROM players").fetchall()]

    computed, _ = compute_state(rows, names)
    new_match = next((m for m in computed if m["id"] == new_id), None)
    return new_match


@app.delete("/matches/{match_id}", status_code=200)
def delete_match(match_id: int):
    with get_db() as db:
        if not db.execute("SELECT id FROM matches WHERE id=?", (match_id,)).fetchone():
            raise HTTPException(404, "Match not found")
        db.execute("DELETE FROM matches WHERE id=?", (match_id,))
    return {"deleted": match_id}


@app.get("/stats")
def get_stats():
    with get_db() as db:
        match_rows = db.execute("SELECT id,date,p1,p2,s1,s2 FROM matches ORDER BY id").fetchall()
        names      = [r["name"] for r in db.execute("SELECT name FROM players ORDER BY name").fetchall()]
    computed, stats = compute_state(match_rows, names)
    return {
        "players": sorted(stats.values(), key=lambda x: -x["elo"]),
        "matches": computed,
    }


@app.get("/export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    with get_db() as db:
        match_rows = db.execute("SELECT id,date,p1,p2,s1,s2 FROM matches ORDER BY id").fetchall()
        names      = [r["name"] for r in db.execute("SELECT name FROM players ORDER BY name").fetchall()]

    computed, stats = compute_state(match_rows, names)

    wb   = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="1B3A6B")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri")

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")

    # Leaderboard sheet
    ws = wb.active
    ws.title = "Leaderboard"
    style_header(ws, ["Rank","Player","ELO","Matches","Wins","Losses","Win %","Peak ELO","Low ELO"])
    for rank, p in enumerate(sorted(stats.values(), key=lambda x: -x["elo"]), 1):
        ws.append([rank, p["name"], p["elo"], p["matches"], p["wins"], p["losses"],
                   round(p["wins"] / p["matches"], 3) if p["matches"] else 0,
                   p["peak"], p["low"]])

    # Match Log sheet
    ws2 = wb.create_sheet("Match Log")
    style_header(ws2, ["ID","Date","Player 1","Player 2","P1 Score","P2 Score",
                        "Valid","Winner","P1 Pre-ELO","P2 Pre-ELO","P1 Post-ELO","P2 Post-ELO"])
    for m in computed:
        ws2.append([m["id"], m["date"], m["p1"], m["p2"], m["s1"], m["s2"],
                    "VALID" if m["valid"] else "INVALID", m["winner"] or "",
                    m["p1pre"], m["p2pre"],
                    m["p1post"] if m["p1post"] is not None else "",
                    m["p2post"] if m["p2post"] is not None else ""])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=squash_elo.xlsx"}
    )
